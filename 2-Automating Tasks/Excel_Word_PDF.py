# author: Allyson Vasquez
# version: June.04.2020
# Notes: Excel, Word, PDF Docs

import openpyxl
import os
from openpyxl import load_workbook

# open spreadsheet & create workbook obj
workbook = openpyxl.load_workbook('example.xlsx')

# workbooks contain multiple sheets - print all sheet names
print(workbook.sheetnames)

# select first available sheet
sheet = workbook.active
print(sheet.title)  # Sheet 1

# Retrieving data from spreadsheet
print(sheet['A1'].value)
print(sheet['B1'].value)
print(sheet['C1'].value)

# Iterating through data
for row in sheet.rows:
    print(row)